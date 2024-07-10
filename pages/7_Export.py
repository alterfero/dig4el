import json
from openpyxl import Workbook
from libs import utils as u
import streamlit as st

#this page generates excel spreadsheets from jsons for users to download.
intro_text = ((((("Each Conversational Questionnaire resides in a separate tab. \nWhen adding transcription in a target language, you can delete the columns corresponding to the pivot language(s) you\'re not using. For each sentence, write the corresponding transcription in the 'target' column. In the 'lexicon' columns, write to the right of each word the corresponding word(s) in the target language, exactly as it appears in the transcription of the sentence. If there is no word or words in the target sentence that are expressing the word in the pivot language, don't write anything. Many words in the lexicon columns won't correspond to anything in the target language (different segmentation, different concepts...). Input what makes sense, discard the rest. If the concept in pivot language is represented by several words in the target language, write all the words separated by three periods. For example, 'not' could be represented by 'ne...pas' in French or 'et...te' in Mwotlap. "+
              "For the sake of readability we did not use too many abreviations, except for personal pronouns, to account for the multiple potential values of, for example, the english 'we'. Personal pronouns are coded 'PPXXXX'. For example, 'PP1SG' is the first person singular, and 'PP2EXCDU' the second person dual exclusive. For possession markers, the same codes are used, preceded by 'POSS'. For example 'POSS PP1PLU' would be the representation of 'our' in 'The leaders in our community'.") +
              "You  can fine-tune dialogs to make them sound more familiar and make more sense. For example : ")+
              "'Village X', X can be replaced by the name of an actual other village, town or city.")+
              "'river', 'lake', or 'forest' for example can become 'sea', 'lagoon', 'field', 'bush' etc. depending on the local environment. Indicate these changes to the comment section of each CQ.  'Engagement day' or 'wedding' etc. can also be replaced if needed")+
              "When using French as the pivot language, the  PP2SG 'vous' marking respect in the patient-doctor dialog can be dropped for a 'tu' if the local practice privileges this form. ")

questionnaires_folder = "./questionnaires"

cqs = u.list_cqs()

target_language = st.selectbox("choose a target language",("marquesan (Nuku Hiva)", "french"))
pivot_language = st.selectbox("choose a pivot language",("english", "french"))

wb = Workbook()
ws_intro = wb.create_sheet("Intro")
ws_intro.cell(row=3, column=2, value = intro_text)

filename = "./recording_template.xlsx"
wb.save(filename)

def create_excel_template(json_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Add headers
    headers = ["counter", "Legacy Index", "English", "Alternate Pivot", "Equivalent in "]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Populate data
    for row, dialog in enumerate(json_data, start=2):
        ws.cell(row=row, column=1, value=dialog['id'])
        ws.cell(row=row, column=2, value=dialog['text'])
        # Leave translation and word relationships columns empty for user input

    filename = "translation_template.xlsx"
    wb.save(filename)
    return filename


